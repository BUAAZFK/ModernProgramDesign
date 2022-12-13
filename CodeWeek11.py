from multiprocessing import Process, Manager, Pipe
import matplotlib.pyplot as plt
from tqdm import tqdm
import pandas as pd
import jieba as jb
import openpyxl
import random
import pickle
import json
import xlwt
import time


def GetData(filepath):
	'''
	filepath: 待读取的文件路径
	return: 元素为新闻内容的list
	'''
	with open(filepath, encoding="utf-8") as f:
		data = json.load(f)
	# data: 一个列表，列表元素为字典，只需用到字典中的content属性
	for i in range(len(data)):
		data[i] = data[i]['content']
	return data

def Clean_Data(datas, stop_words):
	'''
	datas: 词列表
	'''
	for word in datas:
		word.strip()
		if word in stop_words or len(word)==0:
			datas.remove(word)
	return datas

def Map(name, data, freqs, child_conn):
	'''
	name: 进程名称
	data: 文件内容列表
	stop_words: 停用词列表
	child_conn: 管道的一端
	'''
	# 获取程序执行的开始时间
	# start_time = time.time()
	# 分词
	# 获取停用词列表
	print(f'{name} 正在执行……')
	# stop_words = child_conn.recv()
	# print(stop_words[:5])
	cut_texts = []
	print('正在分词并去除停用词')
	for text in tqdm(data):
		cut_texts.append(jb.lcut(text))
	# 词频统计
	freq = {}
	print('正在统计词频')
	for cut_text in tqdm(cut_texts):
		for word in cut_text:
			freq[word] = freq.get(word, 0) + 1
	# end_time = time.time()
	# print(f'{name} 执行了 {end_time-start_time}s …………')
	# 向父进程发送该进程生成的词频字典
	child_conn.send(freq)
	freqs.append(freq)
	# with open('name', 'wb') as f:
	# 	pickle.dump()

def Write_xlwt(filepath, datas):
	'''
	filepath: 写入的文件路径
	datas: 写入的数据,元组或二维列表
	'''
	# 将词频统计结果写入excel文件
	file = xlwt.Workbook() # 创建excel文件
	sheet = file.add_sheet('freq') # 创建sheet表格
	sheet.write(0, 0, 'Word') # 写入header
	sheet.write(0, 1, 'frequency')
	for i in range(len(datas)):
		sheet.write(i+1, 0, datas[i][0])
		sheet.write(i+1, 1, datas[i][1])
	file.save(filepath)

def Write_openpyxl(filepath, datas):
	# 将数据写入excel表格
	workbook = openpyxl.Workbook() 
	sheet = workbook.create_sheet('frequency', index=0) # 创建sheet
	sheet.column_dimensions['A'].width=15 # 设置A列宽度
	sheet.column_dimensions['B'].width=22 # 设置B列宽度
	# 循环写入数据，居中对齐
	for i in range(100000):
		sheet.cell(i+1, 1).value = datas[i][0] # 写入数据
		sheet.cell(i+1, 2).value = datas[i][1] # 写入数据
	workbook.save(filepath) # 保存文件

def Reduce(freqs, filepath):
	'''
	freqs: 所有进程的词频统计字典，是一个列表
	'''
	print('正在合并词频字典')
	total_freq = {}
	for freq in tqdm(freqs):
		for item in list(freq.items()):
			total_freq[item[0]] = total_freq.get(item[0], 0) + item[1]
	freq_tuples = list(total_freq.items()) # 获取词频统计元组
	freq_tuples = sorted(freq_tuples, key = lambda x:x[1], reverse=True)
	# Write_xlwt('./frequency.xls')
	Write_openpyxl(filepath, freq_tuples)
	return total_freq

if __name__ == '__main__':
	# 读取到的数据的长度为1245835
	datas = GetData('./sohu_data.json')
	# 读取停用词列表
	# stop_words = pd.read_table('./stopwords_list.txt', on_bad_lines='skip',quoting=3, header=None).astype(str)
	# stop_words = Manager().list(stop_words[0])
	# stop_words_dict = {}
	# for word in stop_words:
	# 	stop_words_dict[word] = word
	# 	jb.add_word(word)
	'''
	# 设置每个进程处理的数据的数量
	start_index = 0
	end_index = random.randint(300000, 400000)
	# 动态设置进程数量
	p_num = 0
	print("创建进程--------------------------------")
	while start_index != end_index:
		# 这里需要将数据进行分割
		print(f'进程{p_num}: {start_index}-{end_index}')
		p = Process(target=Map, args=('进程 {}'.format(p_num+1), datas[start_index:end_index], freqs))
		start_index = end_index
		end_index = min(start_index+random.randint(300000, 400000), len(datas))
		p_num += 1
		plist.append(p)
	'''
	pnum = 16
	times = []
	for index in range(15, 16):
		print(f'{index+1}个进程执行情况-------------------------------------')
		# 词频字典列表
		# 使用Manager创建的对象在子进程中可以使用
		# freqs = Manager().list()
		freqs = []
		# 使用管道实现通信
		child_conn, parent_conn = Pipe()
		# 向子进程发送停用词列表
		# parent_conn.send(stop_words_dict)
		# 进程列表
		plist=[]
		left = 0
		right = int(len(datas)/(index+1))
		start_time = time.time()
		for i in range(index+1):
			p = Process(target=Map, args=('进程 {}'.format(i+1), datas[left:right], freqs, child_conn))
			left = right
			right += right
			plist.append(p)
		# 启动进程
		for item in plist:
			item.start()
		# 获得子进程的词频列表
		for item in plist:
			freqs.append(parent_conn.recv())
			print(len(freqs))
		# 所有进程都已经启动，使用join进行阻塞，即子进程执行时主进程不执行（所以join的时间是最长的那个子进程执行时间）
		for item in plist:
			item.join()
		end_time = time.time()
		print(f'{index+1}个进程运行耗时{end_time-start_time}s')
		times.append(end_time-start_time)
		total_freq = Reduce(freqs, f'./frequency_{index}.xlsx')
	# x = range(1,pnum+1)
	# plt.plot(x, times)
	# plt.savefig('./timeWaste.png')
	# plt.show()
